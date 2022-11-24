# imports

import os
import pandas as pd
import numpy as np
import math
import openpyxl as xl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, fills
from datetime import datetime
from platform import python_version

ver = python_version()
start_time = datetime.now()
octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}


def rdfile(filename):
    try:
        df = pd.read_excel(f'input/{filename}')
        rows = df.shape[0]
    except Exception as e:
        print("Error in reading Excel file!", e)
        exit()
    return df, rows

def avarage(df):
    try:
        # Calculating avg Values
        u_avg = df['U'].mean()
        v_avg = df['V'].mean()
        w_avg = df['W'].mean()

        # Calculating Average Value of U, V, W
        df.insrt(4, column="U Avg", value="")
        df.insrt(5, column="V Avg", value="")
        df.insrt(6, column="W Avg", value="")

        # Calculating U', V', W' 
        df.insrt(7, column="U'=U - U avg", value="")
        df.insrt(8, column="V'=V - V avg", value="")
        df.insrt(9, column="W'=W - W avg", value="")

        df["U'=U - U avg"] = round(df['U'] - u_avg, 3)
        df["V'=V - V avg"] = round(df['V'] - v_avg, 3)
        df["W'=W - W avg"] = round(df['W'] - w_avg , 3)
        
        df.at[0, 'U Avg'] = round(u_avg, 3)
        df.at[0, 'V Avg'] = round(v_avg, 3)
        df.at[0, 'W Avg'] = round(w_avg, 3)
        
        df['U'] = round(df['U'], 3)
        df['V'] = round(df['V'], 3)
        df['W'] = round(df['W'], 3)
        df['T'] = round(df['T'], 3)
    except:
        print("Error in calculating average!")
        exit()
    return df

def insrt_octant(df):
    try:
        #insrting new column for Octant
        df.insrt(10, column="Octant", value="")
        df.insrt(11, column=" ", value="")
        df.insrt(12, column="", value="")
        df.insrt(13, column="Octant ID", value="")
        df.insrt(14, column="1", value="")
        df.insrt(15, column="-1", value="")
        df.insrt(16, column="2", value="")
        df.insrt(17, column="-2", value="")
        df.insrt(18, column="3", value="")
        df.insrt(19, column="-3", value="")
        df.insrt(20, column="4", value="")
        df.insrt(21, column="-4", value="")

        df.iloc[1, 12] = "Mod "+ str(mod)
        df.at[0, 'Octant ID'] = "Overall Count"
        l=[]
    except:
        print("Error in insrting columns.")
        exit()
    return df

def calculate_octant(df, rows):
    try:
        # Calculating the octant values
        for i in range(0, rows):
            if df.at[i,"U'=U - U avg"] >= 0 and  df.at[i,"V'=V - V avg"] >= 0:
                if df.at[i,"W'=W - W avg"] >= 0:
                  df.at[i, 'Octant'] = 1
                else:
                  df.at[i, 'Octant'] = -1
            elif df.at[i,"U'=U - U avg"] < 0 and  df.at[i,"V'=V - V avg"] >= 0:
                if df.at[i,"W'=W - W avg"] >= 0:
                  df.at[i, 'Octant'] = 2
                else:
                  df.at[i, 'Octant'] = -2
            elif df.at[i,"U'=U - U avg"] < 0 and  df.at[i,"V'=V - V avg"] < 0:
                if df.at[i,"W'=W - W avg"] >= 0:
                  df.at[i, 'Octant'] = 3
                else:
                  df.at[i, 'Octant'] = -3
            elif df.at[i,"U'=U - U avg"] >= 0 and  df.at[i,"V'=V - V avg"] < 0:
                if df.at[i,"W'=W - W avg"] >= 0:
                  df.at[i, 'Octant'] = 4
                else:
                  df.at[i, 'Octant'] = -4
            l.append(df.at[i, 'Octant'])

        df.at[0, "1"] = l.cnt(1)
        df.at[0, "-1"] = l.cnt(-1)
        df.at[0 ,"2"] = l.cnt(2)
        df.at[0 ,"-2"] = l.cnt(-2)
        df.at[0 ,"3"] = l.cnt(3)
        df.at[0 ,"-3"] = l.cnt(-3)
        df.at[0 ,"4"] = l.cnt(4)
        df.at[0 ,"-4"] = l.cnt(-4)
    except:
        print("Error in counting octant values.")
        exit()
    
    try:
        # Splitting list into ranges and finding the count of octant values
        start = 0
        end = len(l)
        step = int(mod)
        idx=1
        total_rows_mod = math.ceil(rows/step)
        total_rows = total_rows_mod
        for i in range(start, end, step):
            x = i
            sub_list = l[x:x+step]
            y = x+step-1
            if y>rows:
                y=rows-1
            df.at[idx ,'Octant ID'] = str(x)+"-"+str(y)
            df.at[idx, '1'] = sub_list.cnt(1)
            df.at[idx, '-1'] = sub_list.cnt(-1)
            df.at[idx, '2'] = sub_list.cnt(2)
            df.at[idx, '-2'] = sub_list.cnt(-2)
            df.at[idx, '3'] = sub_list.cnt(3)
            df.at[idx, '-3'] = sub_list.cnt(-3)
            df.at[idx, '4'] = sub_list.cnt(4)
            df.at[idx, '-4'] = sub_list.cnt(-4)
            idx+=1
    except:
        print("Error in counting octant values for ranges!")
        exit()
    
    try:
        # insrting Rank Columns 
        col_num = 22
        for i in range(1,5):
            header = "Rank Octant "+str(i)
            df.insrt(col_num, column=header, value="")
            col_num+=1
            header = "Rank Octant "+str(-1*i)
            df.insrt(col_num, column=header, value="")
            col_num+=1
        df.insrt(col_num, column="Rank 1 Octant ID", value="")
        col_num+=1
        df.insrt(col_num, column="Rank 1 Octant Name", value="")
        col_num+=1
        
        # Calculating rank for Overall Octant Count
        dict={}
        l=[]
        for i in range(1,5):
            oct_cnt = df.at[0, str(i)]
            dict[oct_cnt] = str(i)
            l.append(oct_cnt)
            oct_cnt = df.at[0, str(-1*i)]
            dict[oct_cnt] = str(-1*i)
            l.append(oct_cnt)
        
        l.sort(reverse=True)
        rank = 1
        df.at[0, "Rank 1 Octant ID"] = dict[l[0]]
        df.at[0, "Rank 1 Octant Name"] = octant_name_id_mapping[dict[l[0]]]
        
        for i in l:
            oct_id = "Rank Octant "+dict[i]
            df.at[0, oct_id] = rank
            rank+=1
        
        # Calculating rank for Mod Octant Count
        rank1=[]
        for idx in range(1, total_rows_mod+1): 
            dict={}
            l=[]
            for i in range(1,5):
                oct_cnt = df.at[idx, str(i)]
                dict[oct_cnt] = str(i)
                l.append(oct_cnt)
                oct_cnt = df.at[idx, str(-1*i)]
                dict[oct_cnt] = str(-1*i)
                l.append(oct_cnt)

            l.sort(reverse=True)
            df.at[idx, "Rank 1 Octant ID"] = dict[l[0]]
            rank1.append(dict[l[0]])
            df.at[idx, "Rank 1 Octant Name"] = octant_name_id_mapping[dict[l[0]]]
            
            rank = 1
            for i in l:
                oct_id = "Rank Octant "+dict[i]
                df.at[idx, oct_id] = rank
                rank+=1
        
        # Count of Rank 1 Mod Values
        idx = total_rows_mod+5
        df.at[idx, 'Rank Octant 4'] = "Octant ID"
        df.at[idx, 'Rank Octant -4'] = "Octant Name"
        df.at[idx, 'Rank 1 Octant ID'] = "Count of Rank 1 Mod Values"
        idx+=1
        for i in range(1,5):
            oct_id = str(i)
            cnt = rank1.cnt(oct_id)
            df.at[idx, 'Rank Octant 4'] = oct_id
            df.at[idx, 'Rank Octant -4'] = octant_name_id_mapping[oct_id]
            df.at[idx, 'Rank 1 Octant ID'] = cnt
            idx+=1
            
            oct_id = str(-1*i)
            cnt = rank1.cnt(oct_id)
            df.at[idx, 'Rank Octant 4'] = oct_id
            df.at[idx, 'Rank Octant -4'] = octant_name_id_mapping[oct_id]
            df.at[idx, 'Rank 1 Octant ID'] = cnt
            idx+=1
            
    except Exception as e:
        print("Error in calculating rank.", e)
    return df

def octant_transition_count(mod, df):
    try:
        # Reading Excel File
        rows = df.shape[0]
        step = mod
        cols = df.shape[1]
        df.insrt(cols, column="                     ", value="")
        cols+=1
    except:
        print("Error in reading Excel file!")
        exit()
    
    try:
        # Overall Transition Count 
        for l in range(2,12):
            blank = ""
            for i in range(1,l+1):
                blank+=" "
            df.insrt(cols, column=blank, value="")
            cols+=1
        blank_dict={}
        bl_len = 4        
        for i in range(1,5):
            blank=""
            for idx in range(0, bl_len):
                blank += " "
            blank_dict[str(i)] = blank
            bl_len+=1
            blank=""
            for idx in range(0, bl_len):
                blank += " "
            blank_dict[str(-1*i)] = blank
            bl_len+=1
        blank_dict['f'] = '  '
        blank_dict['s'] = '   '
        
        idx=0
        df.at[idx, blank_dict['1']] = 'To'
        idx+=1
        df.at[idx, blank_dict['s']] = 'Count'
        for k in range(-4,5):
            if k==0:
                continue
            df.at[idx, blank_dict[str(k)]] =  k
        idx+=1
        df.at[idx, blank_dict['f']] = "From"

        # Creating dataframe df1 to store values
        data=[]
        df1 = pd.DataFrame(data, index=['1','-1','2','-2','3','-3','4','-4'],
                        columns=['1','-1','2','-2','3','-3','4','-4'])

        df1 = df1.fillna(0)  # For filling 0 to df1

        # Calculating values
        for i in range(0,rows-1):
            first = str(df.at[i,'Octant'])
            second = str(df.at[i+1, 'Octant'])
            df1.at[first, second] += 1

        # Adding values to main dataframe
        for k in range (1,5):
            df.at[idx, blank_dict['s']] = str(k)
            for l in range (-4,5):
                if l==0:
                    continue
                df.at[idx, blank_dict[str(l)]] = df1.at[str(k), str(l)]
            idx+=1
            df.at[idx, blank_dict['s']] = str(-1*k)
            for l in range (-4,5):
                if l==0:
                    continue
                df.at[idx, blank_dict[str(l)]] = df1.at[str(-1*k), str(l)]
            idx+=1
    except Exception as e:
        print("Error in calculating Overall Transition Count!", e)
        exit()

    try:
        # Mod Transition Count
        for i in range(0, rows, step):
            lim = i+step
            if lim>=rows:
                lim = rows
            idx+=2
            df.at[idx, blank_dict['s']] = 'Mod Transition Count'
            idx+=1
            df.at[idx, blank_dict['s']] = str(i)+'-'+str(lim-1)
            df.at[idx, blank_dict['1']] = 'To'
            idx+=1
            df.at[idx, blank_dict['s']] = 'Octant #'
            for k in range(-4,5):
                if k==0:
                    continue
                df.at[idx, blank_dict[str(k)]] =  k
            idx+=1
            df.at[idx, blank_dict['f']] = "From"

            data=[]
            df1 = pd.DataFrame(data, index=['1','-1','2','-2','3','-3','4','-4'],
                            columns=['1','-1','2','-2','3','-3','4','-4'])
            df1 = df1.fillna(0)

            # Calculating values
            if lim==rows:
                lim-=1
            for j in range(i,lim):
                first = str(df.at[j,'Octant'])
                second = str(df.at[j+1, 'Octant'])
                df1.at[first, second] += 1

           # Adding values to the main dataframe
            for k in range (1,5):
                df.at[idx, blank_dict['s']] = str(k)
                for l in range (-4,5):
                    if l==0:
                        continue
                    df.at[idx, blank_dict[str(l)]] = df1.at[str(k), str(l)]
                idx+=1
                df.at[idx, blank_dict['s']] = str(-1*k)
                for l in range (-4,5):
                    if l==0:
                        continue
                    df.at[idx, blank_dict[str(l)]] = df1.at[str(-1*k), str(l)]
                idx+=1
    except:
        print("Error in calculating Mod Transition Count!")
        exit()
    
    try:
        # Exporting dataframe to excel
        return df
    except Exception as e:    
        print("Error in exporting to excel.", e)
        exit()

def octant_longest_subsequence_count_with_range(mod, df, filename):
    try:
        # Reading Excel File
        rows = df.shape[0]
        cols = df.shape[1]
        df.insrt(cols, column="                   ", value="")
        cols+=1
    except Exception as e:
        print("Error in reading Excel file!", e)
        exit()
    
    try:
        # Dataframe to store longest sequence and Count
        data=[]
        df1 = pd.DataFrame(data, index=['1','-1','2','-2','3','-3','4','-4'],
                       columns=['Len', 'Count'])
        df1 = df1.fillna(0)
        
        # Dataframe to store time ranges
        df3 = pd.DataFrame(data, columns=['1','-1','2','-2','3','-3','4','-4'])
        
        prev = df.at[0, 'Octant'] 
        df1.at[str(prev), 'Len'] = 1
        cur_len = 1
        ini = df.at[0,'T']
        fin = df.at[0, 'T']

        for idx in range(1,rows):
            cur = df.at[idx, 'Octant']
            if (cur == prev):
                cur_len+=1
            else:
                cur_len = 1
                ini = df.at[idx, 'T']
            fin = df.at[idx, 'T']
            df4 = df3.cnt(axis=0)
            if (cur_len == df1.at[str(cur), 'Len']):
                df1.at[str(cur), 'Count'] += 1                
                df3.at[df4[str(cur)], str(cur)] = ini
                df3.at[df4[str(cur)]+1, str(cur)] = fin
            elif(cur_len > df1.at[str(cur), 'Len']):
                df1.at[str(cur), 'Count'] = 1
                del df3[str(cur)]
                df3.insrt(7, column = str(cur), value="")
                df3.at[0, str(cur)] = ini
                df3.at[1, str(cur)] = fin
            df3.replace('', np.nan, inplace=True)
            df1.at[str(cur), 'Len'] = max(cur_len, df1.at[str(cur), 'Len'])
            prev = cur

        # insrting values in dataframe
        idx = 0
        for i in range(1,5):
            df.at[idx, 'Octant ##'] = str(i)
            df.at[idx, 'Longest Subsequence Length'] = df1.at[str(i), 'Len']
            df.at[idx, 'Count'] = df1.at[str(i), 'Count']
            idx+=1
            df.at[idx, 'Octant ##'] = str(-1*i)
            df.at[idx, 'Longest Subsequence Length'] = df1.at[str(-1*i), 'Len']
            df.at[idx, 'Count'] = df1.at[str(-1*i), 'Count']
            idx+=1
            
        cols=df.shape[1]
        df.insrt(cols, column="                         ", value="")
        # insrting df3 in main dataframe
        idx=0
        for i in range(1,5):
            
            # For positive i
            df.at[idx, 'Octant ###'] = str(i)
            df.at[idx, 'Longest Subsequence Length '] = df1.at[str(i), 'Len']
            df.at[idx, 'Count '] = df1.at[str(i), 'Count']
            idx+=1
            df.at[idx, 'Octant ###'] = "Time"
            df.at[idx, 'Longest Subsequence Length '] = "From"
            df.at[idx, 'Count '] = "To"
            idx+=1
            for index in range(0, len(df3[str(i)]), 2):
                if np.isnan(df3.at[index, str(i)]):
                    break
                df.at[idx, 'Longest Subsequence Length '] = df3.at[index, str(i)]
                df.at[idx, 'Count '] = df3.at[index+1, str(i)]
                idx+=1
            
            ### For negative i
            df.at[idx, 'Octant ###'] = str(-1*i)
            df.at[idx, 'Longest Subsequence Length '] = df1.at[str(-1*i), 'Len']
            df.at[idx, 'Count '] = df1.at[str(-1*i), 'Count']
            idx+=1
            df.at[idx, 'Octant ###'] = "Time"
            df.at[idx, 'Longest Subsequence Length '] = "From"
            df.at[idx, 'Count '] = "To"
            idx+=1
            
            for index in range(0, len(df3[str(-1*i)]), 2):
                if np.isnan(df3.at[index, str(-1*i)]):
                    break
                df.at[idx, 'Longest Subsequence Length '] = df3.at[index, str(-1*i)]
                df.at[idx, 'Count '] = df3.at[index+1, str(-1*i)]  
                idx+=1     
    except Exception as e:
        print("Error in calculating longest sequence.", e)
        exit()
    
    try:
        # Exporting dataframe to excel 
        # df.to_excel(f'output/{filename[0:len(filename)-5]}_vel_octant_analysis_mod_{mod}.xlsx', index=False)
        return df
    except Exception as e:
        print("Error in exporting to Excel file!", e)
        exit()

def main():
    filename = ""
    df, rows = rdfile(f'input/{filename}')
    df = avarage(df)
    df = insrt_octant(df)
    df = calculate_octant(df, rows)
    df = octant_transition_count(mod, df)
    df = octant_longest_subsequence_count_with_range(mod, df, filename)
    path='input/'+filename  
    outPath='output/'+str(filename[0:len(filename)-5])+'_vel_octant_analysis_mod_'+str(mod)+'.xlsx'        
    worksheet = xl.load_workbook(path)
    sheet = worksheet.active
    fill_pattern = PatternFill(patternType="solid",fgColor="FFFF33")
    sheet['L1'].value=""
    sheet['AG1'].value=""
    sheet['AH1'].value=""
    sheet['AR1'].value=""
    tot_r = df.shape[0]
    tot_c = df.shape[1]
    total_rows = math.ceil(tot_r/mod)
    r=0
    for row in sheet.iter_rows(min_row=3, min_col=1, max_row=tot_r+2, max_col=tot_c):
        c=0
        for cell in row: 
            cell.value = df.iat[r, c]
            c+=1
        r+=1
    for row in sheet.iter_rows(min_row=3, min_col=14, max_row=total_rows+4, max_col=32):
        for cell in row: 
            # print(cell.value, end=" ")
            if(cell.value==1):
                cell.fill = fill_pattern

    # Defining border formats 

    thin_border = Border(left=Side(border_style='thin',color='00000000'),
                    right=Side(border_style='thin',color='00000000'),
                    top=Side(border_style='thin',color='00000000'),
                    bottom=Side(border_style='thin',color='00000000')
                    )
    thick_border = Border(left=Side(border_style='thick',color='00000000'),
                right=Side(border_style='thick',color='00000000'),
                top=Side(border_style='thick',color='00000000'),
                bottom=Side(border_style='thick',color='00000000')
                )
                    

    # Defining size of the table 
    col_num=19
    # Location of table 
    row_loc=2
    col_loc=14

    for i in range (row_loc,row_loc+total_rows+2):
        for j in range (col_loc,col_loc+col_num):
            sheet.cell(row=i, column=j).border=thick_border

    for i in range (total_rows+8,total_rows+17):
        for j in range (29,32):
            sheet.cell(row=i, column=j).border=thick_border
    
    x=4
    for n in range(total_rows+1):
        i=0
        for row in sheet.iter_rows(min_row=x, min_col=35, max_row=x+8, max_col=43):
            for cell in row:
                if(cell.value!=None):
                    cell.border = thick_border
        x+=13

    for row in sheet.iter_rows(min_row=2, min_col=45, max_row=10, max_col=47):
        for cell in row:
            cell.border = thick_border

    max_rows = 1
    for ro in range(2, tot_r):
        if str(sheet.cell(row = ro, column = 50).value) == "nan":
            max_rows = ro
            break
    
    for row in sheet.iter_rows(min_row=2, min_col=49, max_row=max_rows-1, max_col=51):
        for cell in row:
            cell.border = thick_border
    
    sheet['A1']=''
    sheet['B1']=''
    sheet['C1']=''
    sheet['D1']=''
    sheet['A2']='T'
    sheet['B2']='U'
    sheet['C2']='V'
    sheet['D2']='W'
    sheet['E2']='U Avg'
    sheet['F2']='V Avg'
    sheet['G2']='W Avg'
    sheet['H2']="U'=U - U avg"
    sheet['I2']="V'=V - V avg"
    sheet['J2']="W'=W - W avg"
    sheet['K2']='Octant'
    sheet['N1']='Overall Octant Count'
    sheet['N2']='Octant ID'
    sheet['O2']='+1'
    sheet['P2']='-1'
    sheet['Q2']='+2'
    sheet['R2']='-2'
    sheet['S2']='+3'
    sheet['T2']='-3'
    sheet['U2']='+4'
    sheet['V2']='-4'
    sheet['W2']='Rank Octant 1'
    sheet['X2']='Rank Octant -1'
    sheet['Y2']='Rank Octant 2'
    sheet['Z2']='Rank Octant -2'
    sheet['AA2']='Rank Octant 3'
    sheet['AB2']='Rank Octant -3'
    sheet['AC2']='Rank Octant 4'
    sheet['AD2']='Rank Octant -4'
    sheet['AE2']='Rank1 Octant ID'
    sheet['AF2']='Rank1 Octant Name'
    sheet['AI1']='Overall Transition Count'
    sheet['AS1']='Longest Subsquence Length'
    sheet['AS2']='Octant ##'
    sheet['AT2']='Longest Subsquence Length'
    sheet['AU2']='Count'
    sheet['AW1']='Longest Subsquence Length with Range'
    sheet['AW2']='Octant ###'
    sheet['AX2']='Longest Subsquence Length'
    sheet['AY2']='Count'
    
    # Saving current worksheet in output file
    worksheet.save(outPath)


if __name__ == '__main__':
    if ver == "3.8.10":
        print("Correct Version Installed")
    else:
        print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")
    mod = 500
    main(mod)
    end_time = datetime.now()
    print('Duration of Program Execution: {}'.format(end_time - start_time))




